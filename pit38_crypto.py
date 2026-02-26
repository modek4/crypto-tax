# =============================================================================
# PIT-38 CRYPTO — Binance — Urząd Skarbowy w Polsce
#
# PRZECZYTAJ MNIE CAŁEGO, ZANIM UŻYJESZ:
# -----------------------------------------------------------------------------
# Art. 17 ust. 1f updof  — definicja odpłatnego zbycia waluty wirtualnej
    # Przez odpłatne zbycie waluty wirtualnej rozumie się wymianę waluty wirtualnej
    # na prawny środek płatniczy, towar, usługę lub prawo majątkowe inne niż waluta
    # wirtualna lub regulowanie innych zobowiązań walutą wirtualną.
# Art. 17 ust. 1g updof  — kwalifikacja jako kapitały pieniężne
    # Przepis ust. 1 pkt 11 stosuje się również do przychodów uzyskanych w ramach
    # prowadzonej działalności gospodarczej, z wyjątkiem działalności, o której mowa
    # w art. 2 ust. 1 pkt 12 ustawy o przeciwdziałaniu praniu pieniędzy oraz finansowaniu
    # terroryzmu, zaliczanej do przychodów z pozarolniczej działalności gospodarczej.
# -----------------------------------------------------------------------------
# Art. 22 ust. 14-16 updof — koszty uzyskania przychodu
# 14.
    # Koszty uzyskania przychodów z tytułu odpłatnego zbycia waluty wirtualnej
    # stanowią udokumentowane wydatki bezpośrednio poniesione na nabycie waluty
    # wirtualnej oraz koszty związane ze zbyciem waluty wirtualnej, w tym
    # udokumentowane wydatki poniesione na rzecz podmiotów, o których mowa w art. 2
    # ust. 1 pkt 12 ustawy o przeciwdziałaniu praniu pieniędzy oraz finansowaniu terroryzmu
# 15.
    # Koszty uzyskania przychodów, o których mowa w ust. 14, są potrącane w tym roku podatkowym,
    # w którym zostały poniesione, z zastrzeżeniem ust. 16.
# 16.
    # Nadwyżka kosztów uzyskania przychodów, o których mowa w ust. 14, nad przychodami
    # z odpłatnego zbycia waluty wirtualnej uzyskanymi w roku podatkowym powiększa koszty
    # uzyskania przychodów z tytułu odpłatnego zbycia waluty wirtualnej
    # poniesione w następnym roku podatkowym.
# -----------------------------------------------------------------------------
# Art. 30b ust. 1a/1b updof — stawka 19%, definicja dochodu
# 1a.
    # Od dochodów uzyskanych z odpłatnego zbycia walut wirtualnych podatek dochodowy
    # wynosi 19% uzyskanego dochodu.
# 1b.
    # Dochodem z odpłatnego zbycia walut wirtualnych jest osiągnięta w roku podatkowym
    # różnica między sumą przychodów uzyskanych z tytułu odpłatnego zbycia walut wirtualnych
    # a kosztami uzyskania przychodów określonymi na podstawie art. 22 ust. 14-16.
# Art. 30b ust. 5d updof  — zakaz łączenia z innymi kapitałami
    # Dochodów z odpłatnego zbycia walut wirtualnych nie łączy się z dochodami opodatkowanymi
    # na zasadach określonych w ust. 1 oraz w art. 27 lub art. 30c.
# Art. 30b ust. 6/6a updof — obowiązek złożenia PIT-38
# 6.
    # Po zakończeniu roku podatkowego podatnik jest obowiązany w zeznaniu, o którym mowa w
    # art. 45 ust. 1a pkt 1, wykazać uzyskane w roku podatkowym dochody, o których mowa w
    # ust. 1 i 1a, i obliczyć należny podatek dochodowy.
# 6a.
    # W zeznaniu, o którym mowa w art. 45 ust. 1a pkt 1, podatnik wykazuje koszty uzyskania
    # przychodów, o których mowa w art. 22 ust. 14-16, także wtedy, gdy w roku podatkowym nie
    # uzyskał przychodów z odpłatnego zbycia walut wirtualnych.
# CAŁOŚĆ PRAWA PODATKOWEGO 2026.01.01:
# https://przepisy.gofin.pl/przepisy,4,16,13,700,,20260101,ustawa-z-dnia-26071991-r-o-podatku-dochodowym-od-osob.html
# -----------------------------------------------------------------------------
# WAŻNE
#
# To narzędzie pomocnicze. Zawsze skonsultuj wynik z doradcą podatkowym.
# Przeczytaj cały kod, komentarze i założenia (zwłaszcza w config.py).
# Sprawdź każdą transakcję, zwłaszcza te oznaczone jako "warning".
# =============================================================================

import pandas as pd
import requests
import time
import os
import sys
import math
from openpyxl import load_workbook
from datetime import datetime, timedelta, timezone
from decimal import Decimal, ROUND_HALF_UP
from config import (
    TARGET_YEAR,
    CSV_FILE,
    CSV_SEPARATOR,
    OUTPUT_FILE,
    DATE_FORMATS,
    CSV_FIELDS,
    CARRIED_COSTS_FROM_PREVIOUS_YEARS,
    FIAT_CURRENCIES,
    STABLECOINS,
    NBP_API_BASE,
    BINANCE_KLINES,
    TRADE_OPS,
    TAXABLE_INCOME_OPS,
    TECHNICAL_OPS,
    BINANCE_DOWNLOAD_URL,
)

# =============================================================================
# KLASA CACHE DLA KURSÓW NBP
# =============================================================================

class NBPRateCache:
    # Pobiera kursy walut z API NBP (Tabela A — średni kurs NBP) do Cache, aby NBP się nie obraziło.
    # Zgodnie z art. 22 ust. 1 updof: kurs z dnia POPRZEDZAJĄCEGO datę transakcji.
    def __init__(self):
        self.cache: dict[str, float] = {}
        self._request_count = 0

    def get_pln_rate(self, currency: str, transaction_datetime: datetime) -> float:
        # Zwraca kurs PLN dla danej waluty.
        # Szuka kursu z DNIA POPRZEDZAJĄCEGO transaction_datetime (zgodnie z prawem).
        # Jeśli to weekend/święto, cofa się maksymalnie 14 dni.
        if currency == "PLN":
            return 1.0

        currency_upper = currency.upper()

        # Data poprzedzająca — art. 22 ust. 1 updof
        prev_day = transaction_datetime.date() - timedelta(days=1)

        for days_back in range(0, 14):
            check_date = prev_day - timedelta(days=days_back)
            date_key = check_date.strftime("%Y-%m-%d")
            cache_key = f"{currency_upper}_{date_key}"

            if cache_key in self.cache:
                return self.cache[cache_key]

            try:
                url = f"{NBP_API_BASE}/{currency_upper.lower()}/{date_key}/"
                resp = requests.get(url, timeout=8)
                self._request_count += 1
                # Żeby API NBP nie wybuchło
                # Zostaw ten czas w spokoju
                time.sleep(0.05)

                if resp.status_code == 200:
                    # Sukces — kurs znaleziony
                    data = resp.json()
                    rate = float(data["rates"][0]["mid"])
                    self.cache[cache_key] = rate
                    return rate
                elif resp.status_code == 404:
                    # Brak notowania w ten dzień (weekend/święto)
                    continue
                elif resp.status_code == 429:
                    # Jak NBP się obrazi, to czekamy dłużej
                    print("\n🟡 NBP rate limit — czekam 5 sekund...")
                    time.sleep(5)
                    continue
            except requests.exceptions.Timeout:
                # Timeout — spróbuj ponownie
                time.sleep(1)
                continue
            except Exception:
                # Inny błąd — logujemy
                continue
        raise ValueError(
            f"🔴 Brak kursu NBP dla '{currency_upper}' w pobliżu daty "
            f"🔴 {transaction_datetime.date()} (sprawdzono 14 dni wstecz). "
            f"🔴 Sprawdź czy waluta jest notowana w tabeli A NBP."
        )

    @property
    def stats(self):
        return f"🟢 Wykonano {self._request_count} zapytań do NBP API"


# =============================================================================
# WYCENA KRYPTO W USD → PLN (dla prowizji i dochodów z earningu)
# =============================================================================

_klines_cache: dict[str, float] = {}

def get_crypto_price_usd(symbol: str, dt: datetime) -> float | None:
    # Pobiera cenę kryptowaluty w USD w danym momencie czasu.
    # Używa 1-godzinnej świecy Binance (najbliższa godzina przed transakcją).
    # Zwraca None jeśli nie można wycenić (token nieznany, para nie istnieje).
    if symbol in STABLECOINS:
        return 1.0
    # Fiat wyceniamy przez NBP, nie przez klines
    if symbol in FIAT_CURRENCIES:
        return None
    # Zaokrąglamy do początku godziny (najbliższa świeca przed transakcją)
    dt_floored = dt.replace(minute=0, second=0, microsecond=0)
    cache_key = f"{symbol}_{dt_floored.strftime('%Y%m%d%H')}"
    # Najpierw sprawdzamy cache, żeby nie robić zbędnych zapytań
    if cache_key in _klines_cache:
        return _klines_cache[cache_key]
    # Próbujemy pary: SYMBOL/USDT, SYMBOL/BTC (jako fallback)
    pairs_to_try = [f"{symbol}USDT", f"{symbol}BTC", f"{symbol}ETH", f"{symbol}BNB"]
    # Jeśli to stablecoin, to próbujemy tylko pary z USDT
    for pair in pairs_to_try:
        try:
            # Binance API używa timestampów w ms
            start_ms = int(dt_floored.replace(tzinfo=timezone.utc).timestamp() * 1000)
            params = {
                "symbol": pair,
                "interval": "1h",
                "startTime": start_ms,
                "limit": 1,
            }
            resp = requests.get(BINANCE_KLINES, params=params, timeout=8)
            # Żeby API Binance nie wybuchło
            # Zostaw ten czas w spokoju
            time.sleep(0.05)
            if resp.status_code == 200:
                data = resp.json()
                if data:
                    # Cena zamknięcia świecy to nasza wycena (najbliższa godzina przed transakcją)
                    close_price = float(data[0][4])
                    # Jeśli para była np. SYMBOL/BTC, mnożymy przez cenę BTC
                    if pair.endswith("BTC"):
                        btc_price = get_crypto_price_usd("BTC", dt)
                        if btc_price:
                            close_price *= btc_price
                        else:
                            continue
                    # Jeśli para była np. SYMBOL/ETH, mnożymy przez cenę ETH
                    elif pair.endswith("ETH"):
                        eth_price = get_crypto_price_usd("ETH", dt)
                        if eth_price:
                            close_price *= eth_price
                        else:
                            continue
                    # Jeśli para była np. SYMBOL/BNB, mnożymy przez cenę BNB
                    elif pair.endswith("BNB"):
                        bnb_price = get_crypto_price_usd("BNB", dt)
                        if bnb_price:
                            close_price *= bnb_price
                        else:
                            continue
                    _klines_cache[cache_key] = close_price
                    return close_price
        except Exception:
            continue
    # Nie udało się wycenić — para nie istnieje lub token jest nieznany
    return None


# =============================================================================
# PARSOWANIE DAT
# =============================================================================

def parse_datetime_column(series: pd.Series) -> pd.Series:
    # Próbuje sparsować kolumnę dat w różnych formatach.
    # Zwraca pd.Series z wartościami datetime lub NaT dla nieparsowanych.
    # Binance może mieć różne formaty daty w zależności od eksportu, więc próbujemy kilku.
    for fmt in DATE_FORMATS:
        try:
            parsed = pd.to_datetime(series, format=fmt)
            # Sprawdzamy, czy parsowanie się powiodło — min 80%
            if parsed.notna().sum() > len(series) * 0.8:
                return parsed
        except Exception:
            continue
    # Jeśli żaden format nie zadziałał, próbujemy bez określonego formatu
    return pd.to_datetime(series, errors="coerce", dayfirst=True)


# =============================================================================
# WALIDACJA I WCZYTYWANIE CSV
# =============================================================================

def load_and_validate_csv(filepath: str, sep: str) -> pd.DataFrame | None:
    # Wczytuje plik CSV z Binance, normalizuje nagłówki,
    # Waliduje czy wymagane kolumny są obecne.
    # Zwraca DataFrame lub None jeśli wystąpił błąd krytyczny (np. brak pliku, brak kolumn).
    if not os.path.exists(filepath):
        print(f"🔴 Nie znaleziono pliku '{filepath}'")
        print(f"🔴 Pobierz go z: '{BINANCE_DOWNLOAD_URL}'")
        print(f"🔴 Upewnij się, że eksportujesz 'Generate all statements' z Download Center Binance.")
        return None
    # Próbujemy automatycznie wykryć separator
    # W config.py (CSV_SEPARATOR) jest ustawiony domyślny separator,
    # Ale Binance może używać różnych (przecinek, średnik, tabulator)
    separators = [sep, ",", ";", "\t"]
    df = None
    for s in separators:
        try:
            tmp = pd.read_csv(filepath, sep=s, encoding="utf-8-sig", engine="python", nrows=5)
            if len(tmp.columns) >= 4:
                df = pd.read_csv(filepath, sep=s, encoding="utf-8-sig", engine="python")
                print(f"🟢 Plik wczytany, separator: '{s}', wierszy: {len(df)}")
                break
        except Exception:
            continue
    if df is None:
        print(f"🔴 Nie udało się wczytać pliku CSV. Sprawdź format i kodowanie (UTF-8).")
        return None
    # Normalizacja nagłówków
    df = df.loc[:, ~df.columns.str.contains("^Unnamed")]
    df.columns = df.columns.str.strip()
    # Spróbuj dostosować nazwy kolumn — obsługa polskich i angielskich nagłówków
    # W config.py (CSV_FIELDS) definiujemy mapowanie, ale tutaj próbujemy automatycznie dopasować
    # Kolumny do naszych standardowych nazw (UTC_Time, Operation, Coin, Change, Account, Remark)
    col_aliases = {
        "UTC_Time": ["UTC_Time", "Czas", "Time", "Date"],
        "Operation": ["Operation", "Operacja", "Type"],
        "Coin":      ["Coin", "Moneta", "Asset", "Currency"],
        "Change":    ["Change", "Zmień", "Amount", "Quantity"],
        "Account":   ["Account", "Konto"],
        "Remark":    ["Remark", "Uwagi", "Note"],
    }
    rename_map = {}
    # Dla każdej kanonicznej nazwy kolumny sprawdzamy, czy któraś z jej aliasów jest obecna w DataFrame
    for canonical, aliases in col_aliases.items():
        for alias in aliases:
            if alias in df.columns and canonical not in df.columns:
                rename_map[alias] = canonical
                break
    if rename_map:
        df = df.rename(columns=rename_map)
        print(f"🔵 Przemianowano kolumny: {rename_map}")
    # Walidacja wymaganych kolumn
    required = ["UTC_Time", "Operation", "Coin", "Change", "Account"]
    missing = [c for c in required if c not in df.columns]
    if missing:
        print(f"🔴 Brakujące kolumny: {missing}")
        print(f"🔴 Dostępne kolumny: {list(df.columns)}")
        print(f"🔴 Upewnij się, że pobierasz 'Generate all statements' z Download Center Binance.")
        return None
    return df


# =============================================================================
# GŁÓWNA FUNKCJA PRZETWARZAJĄCA
# =============================================================================

def process_transaction(row, nbp: NBPRateCache) -> tuple[str, dict | None]:
    # Klasyfikuje pojedynczą transakcję.
    # Zwraca krotkę: (kategoria, dane_słownikowe)
    # Kategorie: "revenue", "cost", "income", "ignored", "warning"
    dt: datetime        = row["_dt"]
    moneta: str         = str(row["Coin"]).upper().strip()
    typ: str            = str(row["Operation"]).strip()
    change_raw: float   = float(str(row["Change"]).replace(",", "."))
    amount: float       = abs(change_raw)
    date_str: str       = dt.strftime("%Y-%m-%d")
    # Podstawowe info wspólne dla wszystkich kategorii
    base_info = {
        "Data":        dt.strftime("%Y-%m-%d %H:%M:%S"),
        "Operacja":    typ,
        "Moneta":      moneta,
        "Ilość":       change_raw,
        "Konto":       row.get("Account", "Spot"),
    }
    # Flagi pomocnicze
    is_fiat        = moneta in FIAT_CURRENCIES
    is_stable      = moneta in STABLECOINS
    is_outflow     = change_raw < 0
    is_inflow      = change_raw > 0

    # ------------------------------------------------------------------
    # 1. OPERACJE HANDLOWE (Buy/Sell/Convert)
    # ------------------------------------------------------------------
    if typ in TRADE_OPS:
        # 1a. KOSZT NABYCIA: wydajemy FIAT → kupujemy krypto (art. 22 ust. 14 updof)
        if is_fiat and is_outflow:
            rate = nbp.get_pln_rate(moneta, dt)
            pln  = round(amount * rate, 6)
            return "cost", {**base_info,
                "PLN":          pln,
                "Kurs_NBP":     rate,
                "Typ":          f"KOSZT NABYCIA — {typ}",
                "Podstawa":     "art. 22 ust. 14 updof"}
        # 1b. PRZYCHÓD: otrzymujemy FIAT za krypto (art. 17 ust. 1f updof)
        elif is_fiat and is_inflow:
            rate = nbp.get_pln_rate(moneta, dt)
            pln  = round(amount * rate, 6)
            return "revenue", {**base_info,
                "PLN":          pln,
                "Kurs_NBP":     rate,
                "Typ":          f"PRZYCHÓD ZE SPRZEDAŻY — {typ}",
                "Podstawa":     "art. 17 ust. 1f updof"}
        # 1c. Krypto → Krypto lub Krypto → Stablecoin: NEUTRALNE
        elif not is_fiat:
            reason = (
                "Wymiana Krypto→Stablecoin (neutralna, stanowisko KIS 2024/2025)"
                if is_stable else
                "Wymiana Krypto→Krypto (neutralna, art. 17 ust. 1f updof)"
            )
            return "ignored", {**base_info, "Powód": reason}

    # ------------------------------------------------------------------
    # 2. PROWIZJE TRANSAKCYJNE (art. 22 ust. 14 updof — koszt)
    # ------------------------------------------------------------------
    elif typ in {"Transaction Fee", "Fee", "Trading Fee"}:
        if is_fiat:
            rate = nbp.get_pln_rate(moneta, dt)
            pln  = round(amount * rate, 6)
            return "cost", {**base_info,
                "PLN":      pln,
                "Kurs_NBP": rate,
                "Typ":      f"PROWIZJA ({moneta})",
                "Podstawa": "art. 22 ust. 14 updof"}
        else:
            # Prowizja w krypto — wyceniamy przez USD→PLN
            usd_price = get_crypto_price_usd(moneta, dt)
            if usd_price is not None:
                usd_rate = nbp.get_pln_rate("USD", dt)
                pln = round(amount * usd_price * usd_rate, 6)
                return "cost", {**base_info,
                    "PLN":          pln,
                    "Kurs_USD_NBP": usd_rate,
                    "Cena_USD":     usd_price,
                    "Typ":          f"PROWIZJA KRYPTO ({moneta}→USD→PLN)",
                    "Podstawa":     "art. 22 ust. 14 updof"}
            else:
                return "warning", {**base_info,
                    "Powód": f"Prowizja w {moneta} — nie udało się wycenić (brak pary na Binance). "
                             f"Wyceń ręcznie i dodaj do kosztów."}

    # ------------------------------------------------------------------
    # 3. PRZYCHODY Z EARNINGU (staking, Launchpool, airdrop, odsetki)
    #    art. 17 ust. 1f updof + interpretacje KIS dot. staking/airdrop
    #    Wartość w dniu otrzymania = przychód + koszt nabycia przy sprzedaży
    # ------------------------------------------------------------------
    elif typ in TAXABLE_INCOME_OPS and is_inflow:
        usd_price = get_crypto_price_usd(moneta, dt)
        if usd_price is not None:
            usd_rate = nbp.get_pln_rate("USD", dt)
            pln = round(amount * usd_price * usd_rate, 6)
            return "income", {**base_info,
                "PLN":          pln,
                "Kurs_USD_NBP": usd_rate,
                "Cena_USD":     usd_price,
                "Typ":          f"PRZYCHÓD EARN/STAKING — {typ}",
                "Podstawa":     "art. 17 ust. 1f updof — wartość rynkowa w dniu otrzymania",
                "UWAGA":        "Ta wartość PLN = Twój koszt nabycia przy późniejszej sprzedaży tej kryptowaluty!"}
        else:
            return "warning", {**base_info,
                "Powód": f"Dochód z {typ} w {moneta} — nie udało się automatycznie wycenić. "
                         f"Wyceń ręcznie (cena {moneta} w {date_str}) i dodaj do PRZYCHODÓW!"}

    # ------------------------------------------------------------------
    # 4. KONWERSJA PYŁU (Small Assets Exchange → BNB)
    #    Neutralna podatkowo w Polsce (krypto→krypto), ale flaga DO WERYFIKACJI
    # ------------------------------------------------------------------
    elif "Small assets exchange" in typ or "Small Assets Exchange" in typ:
        return "ignored", {**base_info,
            "Powód": "Konwersja pyłu → BNB (krypto→krypto, neutralna). "
                     "Jeśli pył był wymieniany na FIAT — wymaga ręcznej korekty!"}

    # ------------------------------------------------------------------
    # 5. TRANSFERY WEWNĘTRZNE I OPERACJE TECHNICZNE
    # ------------------------------------------------------------------
    elif typ in TECHNICAL_OPS:
        return "ignored", {**base_info, "Powód": f"Operacja techniczna ({typ}) — brak skutku podatkowego"}

    # ------------------------------------------------------------------
    # 6. WPŁATY I WYPŁATY FIAT
    # ------------------------------------------------------------------
    elif is_fiat and typ in {"Deposit", "Fiat Deposit"}:
        return "ignored", {**base_info, "Powód": "Wpłata własnych środków fiducjarnych"}
    elif is_fiat and typ in {"Withdraw", "Fiat Withdraw"}:
        return "ignored", {**base_info, "Powód": "Wypłata środków na konto bankowe"}

    # ------------------------------------------------------------------
    # 7. INNE — wymagają ręcznej weryfikacji
    # ------------------------------------------------------------------
    else:
        category = "warning"
        msg = f"Nieznana operacja: '{typ}' dla {moneta}. Wymaga ręcznej klasyfikacji podatkowej!"
        return category, {**base_info, "Powód": msg}

# =============================================================================
# ZAOKRĄGLANIE ZGODNE Z POLSKIM PRAWEM
# Podstawa opodatkowania — pełne złote, podatek — pełne złote (art. 30b updof)
# =============================================================================

def round_pln(value: float) -> int:
    # Zaokrągla do pełnych złotych metodą half-up (Decimal dla precyzji).
    return int(Decimal(str(value)).quantize(Decimal("1"), rounding=ROUND_HALF_UP))

# =============================================================================
# WYKRYWANIE NIESTANDARDOWYCH KONT (Futures, Margin)
# =============================================================================

def warn_non_spot_accounts(df: pd.DataFrame):
    # Ostrzega o transakcjach na kontach innych niż Spot.
    all_accounts = df["Account"].dropna().unique()
    non_spot = [a for a in all_accounts if str(a).strip().lower() != "spot"]
    if non_spot:
        print(f"\n{'='*70}")
        print(f"🟡 Wykryto transakcje na kontach: {non_spot}")
        print(f"🟡 Kontrakty Futures i Margin NIE są obsługiwane przez ten skrypt.")
        print(f"🟡 Wymagają osobnej analizy prawnej (brak jednoznacznych interpretacji).")
        print(f"🟡 Transakcje z tych kont zostały POMINIĘTE.")
        print(f"{'='*70}\n")

# =============================================================================
# AUTOFIT KOLUMN W EXCELU
# =============================================================================

def autofit_columns(sheet):
    for column_cells in sheet.columns:
        max_length = 0
        column = column_cells[0].column_letter
        for cell in column_cells:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        sheet.column_dimensions[column].width = adjusted_width

# =============================================================================
# GŁÓWNA FUNKCJA
# =============================================================================

def main():
    print("=" * 70)
    print(f"  PIT-38 CRYPTO — Binance — Rok podatkowy: {TARGET_YEAR}")
    print("=" * 70)
    # Wczytaj i zwaliduj CSV
    # load_and_validate_csv zwraca DataFrame lub None jeśli wystąpił
    # błąd krytyczny (np. brak pliku, brak kolumn)
    df = load_and_validate_csv(CSV_FILE, CSV_SEPARATOR)
    if df is None:
        return
    # Parsuj daty
    # parse_datetime_column zwraca pd.Series z datetime lub NaT
    # dla nieparsowanych, więc dodajemy tymczasową kolumnę "_dt"
    df["_dt"] = parse_datetime_column(df["UTC_Time"])
    unparsed = df["_dt"].isna().sum()
    if unparsed > 0:
        print(f"🟡 Nie udało się sparsować {unparsed} dat — te wiersze zostaną pominięte.")
    df = df.dropna(subset=["_dt"])
    # Ostrzeż o kontach non-Spot
    # warn_non_spot_accounts nie przerywa działania, ale informuje użytkownika
    # o potencjalnych problemach z danymi
    warn_non_spot_accounts(df)
    # Filtruj rok + konto Spot
    df_year = df[(df["_dt"].dt.year == TARGET_YEAR) & (df["Account"].str.strip() == "Spot")].copy()
    total = len(df_year)
    if total == 0:
        print(f"🔴 Brak transakcji Spot dla roku {TARGET_YEAR}. Sprawdź plik CSV.")
        return
    print(f"\n🔵 Znaleziono {total} transakcji Spot dla roku {TARGET_YEAR}.")
    print(f"🔵 Pobieranie kursów NBP i cen krypto (może chwilę potrwać)...\n")
    # Sortuj chronologicznie (ważne dla FIFO)
    # NBP wymaga kursu z dnia poprzedzającego transakcję, więc kolejność ma znaczenie
    df_year = df_year.sort_values("_dt").reset_index(drop=True)
    # Inicjalizacja
    # Cache dla kursów NBP
    nbp = NBPRateCache()
    # Listy do przechowywania danych dla poszczególnych kategorii
    revenues: list[dict] = []    # przychody ze sprzedaży → Pole 34 PIT-38
    costs: list[dict]    = []    # koszty uzyskania przychodu → Pole 35 PIT-38
    incomes: list[dict]  = []    # przychody z earn/staking → Pole 34 PIT-38
    warnings: list[dict] = []    # wymagają ręcznej weryfikacji
    ignored: list[dict]  = []    # neutralne podatkowo
    # Licznik błędów do statystyk końcowych
    errors_count = 0
    # ------------------------------------------------------------------
    # Główna pętla
    # ------------------------------------------------------------------
    for idx, row in df_year.iterrows():
        progress = idx + 1
        icon = "🔵🔵🔵🔵🔵"
        if progress <= total * 0.25:
            icon = "🟢🔵🔵🔵🔵"
        elif progress <= total * 0.5:
            icon = "🟢🟢🔵🔵🔵"
        elif progress <= total * 0.75:
            icon = "🟢🟢🟢🔵🔵"
        elif progress < total:
            icon = "🟢🟢🟢🟢🔵"
        elif progress == total:
            icon = "🟢🟢🟢🟢🟢"
        if progress % 50 == 0 or progress == total:
            print(f"{icon} {progress}/{total} transakcji...", end="\r")
        try:
            # process_transaction zwraca krotkę: (kategoria, dane_słownikowe)
            category, data = process_transaction(row, nbp)
            if category == "revenue":
                revenues.append(data)
            elif category == "cost":
                costs.append(data)
            elif category == "income":
                incomes.append(data)
            elif category == "warning":
                warnings.append(data)
            else:
                ignored.append(data)
        # Każdy błąd podczas przetwarzania transakcji jest łapany, liczony i logowany do statystyk końcowych
        except ValueError as e:
            errors_count += 1
            ignored.append({
                "Data":     row["_dt"].strftime("%Y-%m-%d %H:%M:%S"),
                "Operacja": str(row.get("Operation", "?")),
                "Moneta":   str(row.get("Coin", "?")),
                "Ilość":    row.get("Change", 0),
                "Konto":    row.get("Account", "?"),
                "Powód":    f"🔴 Błąd wartości: {e} — SPRAWDŹ RĘCZNIE",
            })
        except Exception as e:
            errors_count += 1
            ignored.append({
                "Data":     row["_dt"].strftime("%Y-%m-%d %H:%M:%S") if "_dt" in row else "?",
                "Operacja": str(row.get("Operation", "?")),
                "Moneta":   str(row.get("Coin", "?")),
                "Ilość":    row.get("Change", 0),
                "Konto":    row.get("Account", "?"),
                "Powód":    f"🔴 Nieoczekiwany błąd: {e} — SPRAWDŹ RĘCZNIE",
            })
    # Obliczenia podatkowe
    total_rev_sale    = sum(r["PLN"] for r in revenues)
    total_rev_earn    = sum(i["PLN"] for i in incomes)
    # Pole 34 PIT-38
    total_rev         = total_rev_sale + total_rev_earn
    total_cos_current = sum(c["PLN"] for c in costs)
    # Pole 35 PIT-38
    total_cos         = total_cos_current + CARRIED_COSTS_FROM_PREVIOUS_YEARS
    # Dochód do opodatkowania
    dochod         = max(0.0, total_rev - total_cos)
    # Nadwyżka kosztów → przeniesienie na kolejny rok (art. 22 ust. 16 updof)
    nadwyzka_kosz  = max(0.0, total_cos - total_rev)
    # Zaokrąglenia zgodne z polskim prawem (art. 30b ust. 1a updof)
    podstawa_pln   = round_pln(dochod)
    # Podatek do zapłaty 19% od podstawy opodatkowania (art. 30b ust. 1a updof)
    podatek_pln    = round_pln(podstawa_pln * 0.19)
    # ------------------------------------------------------------------
    # Zapis do Excela
    # ------------------------------------------------------------------
    print(f"\n\n🔵 Zapisuję wyniki do: {OUTPUT_FILE}")
    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        # Arkusz PODSUMOWANIE
        summary_data = [
            ("─" * 50, ""),
            ("POLE 34 — PRZYCHODY (art. 17 ust. 1f updof)", ""),
            ("  Przychody ze sprzedaży krypto→fiat",         round(total_rev_sale, 2)),
            ("  Przychody z earn/staking/airdrop",            round(total_rev_earn, 2)),
            ("  RAZEM Pole 34",                               round(total_rev, 2)),
            ("─" * 50, ""),
            ("POLE 35 — KOSZTY (art. 22 ust. 14-16 updof)", ""),
            (f"  Koszty poniesione w {TARGET_YEAR}",          round(total_cos_current, 2)),
            (f"  Nadwyżka kosztów przeniesiona z lat poprz.", round(CARRIED_COSTS_FROM_PREVIOUS_YEARS, 2)),
            ("  RAZEM Pole 35",                               round(total_cos, 2)),
            ("─" * 50, ""),
            ("DOCHÓD (Pole 34 - Pole 35)",                    round(dochod, 2)),
            ("Podstawa opodatkowania [pełne PLN]",            podstawa_pln),
            ("PODATEK DO ZAPŁATY 19% [pełne PLN]",            podatek_pln),
            ("─" * 50, ""),
            ("NADWYŻKA KOSZTÓW → NASTĘPNY ROK",              round(nadwyzka_kosz, 2)),
            ("  ↑ Wpisz tę kwotę jako CARRIED_COSTS_FROM_PREVIOUS_YEARS", ""),
            ("  ↑ w rozliczeniu za rok następny (art. 22 ust. 16 updof)", ""),
            ("─" * 50, ""),
            ("STATYSTYKI", ""),
            ("  Transakcji przetworzonych",                   total),
            ("  Przychodów (wierszy)",                        len(revenues)),
            ("  Kosztów (wierszy)",                           len(costs)),
            ("  Dochodów Earn/Staking (wierszy)",             len(incomes)),
            ("  Ostrzeżeń do weryfikacji",                    len(warnings)),
            ("  Neutralnych/ignorowanych",                    len(ignored)),
            ("  Błędów przetwarzania",                        errors_count),
        ]
        pd.DataFrame(summary_data, columns=["Opis", "Wartość PLN"]).to_excel(
            writer, sheet_name="PODSUMOWANIE", index=False)
        # Arkusz PRZYCHODY (Pole 34)
        if revenues:
            pd.DataFrame(revenues).to_excel(writer, sheet_name="POLE_34_PRZYCHODY", index=False)
        # Arkusz KOSZTY (Pole 35)
        if costs:
            pd.DataFrame(costs).to_excel(writer, sheet_name="POLE_35_KOSZTY", index=False)
        # Arkusz EARN / STAKING (część Pola 34)
        if incomes:
            pd.DataFrame(incomes).to_excel(writer, sheet_name="EARN_STAKING_POLE34", index=False)
        # Arkusz OSTRZEŻENIA — WYMAGAJĄ RĘCZNEJ WERYFIKACJI
        if warnings:
            pd.DataFrame(warnings).to_excel(writer, sheet_name="WERYFIKACJA_RĘCZNA", index=False)
        # Arkusz IGNOROWANE (neutralne podatkowo)
        if ignored:
            pd.DataFrame(ignored).to_excel(writer, sheet_name="IGNOROWANE", index=False)
    # Formatuj arkusz Excel (szerokości kolumn, style) — tylko jeśli są dane
    if os.path.exists(OUTPUT_FILE):
        wb = load_workbook(OUTPUT_FILE)
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            autofit_columns(sheet)
        wb.save(OUTPUT_FILE)
    else:
        print(f"🔴 Nie udało się znaleźć wygenerowanego pliku '{OUTPUT_FILE}' — sprawdź błędy zapisu.")
    # Raport końcowy
    print("\n" + "=" * 70)
    print(f"🟢 GOTOWE — Wyniki: {OUTPUT_FILE}")
    print("=" * 70)
    print(f"\n🔵 ZESTAWIENIE PIT-38 ZA ROK {TARGET_YEAR}")
    print(f"  ─────────────────────────────────────────────────────")
    print(f"    POLE 34  Przychody:        {total_rev:>12.2f} PLN")
    print(f"      w tym ze sprzedaży:      {total_rev_sale:>12.2f} PLN")
    print(f"      w tym Earn/Staking:      {total_rev_earn:>12.2f} PLN")
    print(f"    POLE 35  Koszty:           {total_cos:>12.2f} PLN")
    print(f"      w tym bieżące:           {total_cos_current:>12.2f} PLN")
    print(f"      w tym poprzednie lata:   {CARRIED_COSTS_FROM_PREVIOUS_YEARS:>12.2f} PLN")
    print(f"  ─────────────────────────────────────────────────────")
    print(f"    Dochód:                    {dochod:>12.2f} PLN")
    print(f"    Podstawa [pełne PLN]:      {podstawa_pln:>12} PLN")
    print(f"    PODATEK 19%:               {podatek_pln:>12} PLN")
    print(f"  ─────────────────────────────────────────────────────")
    print(f"    Nadwyżka→nast. rok:        {nadwyzka_kosz:>12.2f} PLN")
    print(f"  ─────────────────────────────────────────────────────")

    if warnings:
        print(f"\n")
        print(f"     🟡 {len(warnings)} TRANSAKCJI WYMAGA RĘCZNEJ WERYFIKACJI!")
        print(f"     🟡 Otwórz arkusz '🟡 WERYFIKACJA_RĘCZNA' i skonsultuj z doradcą.")

    if errors_count:
        print(f"\n")
        print(f"     🔴 {errors_count} błędów podczas przetwarzania.")
        print(f"     🔴 Sprawdź arkusz 'IGNOROWANE' — wiersze z flagą 'Błąd'.")

    if nadwyzka_kosz > 0:
        print(f"\n")
        print(f"     🔵 Nadwyżka kosztów = {nadwyzka_kosz:.2f} PLN")
        print(f"     🔵 Wpisz tę wartość jako CARRIED_COSTS_FROM_PREVIOUS_YEARS")
        print(f"     🔵 w rozliczeniu za rok {TARGET_YEAR + 1} (art. 22 ust. 16 updof).")

    print(f"\n")
    print(f"🟡 Ten skrypt to narzędzie pomocnicze.")
    print(f"{nbp.stats}")
    print(f"🟡 Zawsze zweryfikuj wynik z doradcą podatkowym przed złożeniem PIT-38.")

if __name__ == "__main__":
    main()
